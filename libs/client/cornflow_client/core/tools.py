"""

"""

# Full imports
import io
import json
import pickle

# Partial imports
from openpyxl import Workbook
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from pytups import OrderSet
from typing import Optional


def new_set(seq):
    """
    :param seq: a (hopefully unique) list of elements (tuples, strings, etc.)
    Returns a new ordered set
    """
    return OrderSet(seq)


def load_json(path):
    with open(path) as json_file:
        file = json.load(json_file)
    return file


def save_json(data, path):
    with open(path, "w") as outfile:
        json.dump(data, outfile)


def copy(dictionary):
    return pickle.loads(pickle.dumps(dictionary, -1))


# Frontend formatting palettes, shared by the inline formatter below.
# Keyed by style set: "master_tables" (blue header theme) and
# "instance_solution" (gray banded-row theme).
_FORMAT_STYLES = {
    "master_tables": {
        "header": {
            "fill": PatternFill(
                start_color="4A90E2", end_color="4A90E2", fill_type="solid"
            ),
            "font": Font(bold=True, color="FFFFFF"),
            "border": Border(
                left=Side(style="thin", color="FFFFFF"),
                right=Side(style="thin", color="FFFFFF"),
                top=Side(style="thin", color="FFFFFF"),
                bottom=Side(style="thin", color="FFFFFF"),
            ),
        },
        "even_row": {
            "fill": PatternFill(
                start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
            ),
            "font": Font(bold=False, color="000000"),
            "border": Border(
                left=Side(style="thin", color="E1E5E9"),
                right=Side(style="thin", color="E1E5E9"),
                top=Side(style="thin", color="E1E5E9"),
                bottom=Side(style="thin", color="E1E5E9"),
            ),
        },
        "odd_row": {
            "fill": PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            ),
            "font": Font(bold=False, color="000000"),
            "border": Border(
                left=Side(style="thin", color="E1E5E9"),
                right=Side(style="thin", color="E1E5E9"),
                top=Side(style="thin", color="E1E5E9"),
                bottom=Side(style="thin", color="E1E5E9"),
            ),
        },
    },
    "instance_solution": {
        "header": {
            "fill": PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            ),
            "font": Font(bold=True, color="000000"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        },
        "even_row": {
            "fill": PatternFill(
                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
            ),
            "font": Font(bold=False, color="000000"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        },
        "odd_row": {
            "fill": PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            ),
            "font": Font(bold=False, color="000000"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        },
    },
}


def _add_frontend_formatting(ws, n_rows, n_cols, style_set="master_tables"):
    """
    Applies the frontend format to a worksheet.

    :param style_set: which palette in _FORMAT_STYLES to use, either
        "master_tables" (blue header) or "instance_solution" (gray banded).
    """
    if n_rows == 0 or n_cols == 0:
        return
    styles = _FORMAT_STYLES[style_set]
    max_cell = f"{get_column_letter(n_cols)}{n_rows}"
    for key, formula in (
        ("header", ["ROW()=1"]),
        ("even_row", ["AND(MOD(ROW(),2)=0, ROW()<>1)"]),
        ("odd_row", ["AND(MOD(ROW(),2)=1, ROW()<>1)"]),
    ):
        style = styles[key]
        dxf = DifferentialStyle(
            fill=style["fill"], font=style["font"], border=style["border"]
        )
        ws.conditional_formatting.add(
            f"A1:{max_cell}",
            Rule(type="expression", dxf=dxf, formula=formula),
        )


def _jsonify(value):
    if isinstance(value, (str, float, int, bool)) or value is None:
        return value
    return json.dumps(value)


def to_excel_memory_file(
    data: dict, is_instance_solution: bool = True
) -> Optional[io.BytesIO]:
    """
    Converts a json into a formatted Excel memory file.
    Uses openpyxl's write_only mode and applies the frontend formatting inline,
    in the same single pass that streams the rows.

    :param data: a json dictionary
    :param is_instance_solution: if True, use the "instance_solution" style set
        (gray banded rows); otherwise use the default "master_tables" style set
        (blue header).
    :return: Formatted Excel file in memory as a BytesIO object
    """
    if not data:
        return None
    style_set = "instance_solution" if is_instance_solution else "master_tables"
    wb = Workbook(write_only=True)
    used_table_names = {}
    max_length = 31
    for table_name, table_data in data.items():
        # Handle the case where sheet names are too long.
        truncated_name = table_name
        if len(table_name) > max_length:
            truncated_name = table_name[:max_length]
        truncated_name_w_suffix = truncated_name
        if truncated_name in used_table_names:
            current_suffix = used_table_names[truncated_name] + 1
            truncated_name_w_suffix = (
                f"{truncated_name[:max_length - 3]}_{current_suffix}"
            )
        used_table_names[truncated_name] = 0
        truncated_name = truncated_name_w_suffix

        # If table is a configuration table {param_1: 1, param_2: 5},
        #   transform it to a list of dicts name/value: [{'name': 'param_1', 'value': 1}, ...]
        if isinstance(table_data, dict):
            table_data = [{"name": k, "value": v} for k, v in table_data.items()]

        ws = wb.create_sheet(title=truncated_name)
        if len(table_data) == 0:
            continue

        # Infer table format: if the items are primitives,
        #   treat as a single "value" column; if dicts, use keys as columns.
        if isinstance(table_data[0], (str, float, int, bool)):
            table_data = [{"value": v} for v in table_data]

        # Build the headers
        headers = list(table_data[0].keys())
        for row in table_data[1:]:
            for key in row:
                if key not in headers:
                    headers.append(key)

        # Measure max string width per column over the header + first 1000 data
        # rows, to set the column widths.
        WIDTH_SAMPLE = min(len(table_data), 1000)
        col_widths = [
            max(
                len(str(h)),
                *[len(str(row.get(h))) for row in table_data[:WIDTH_SAMPLE]],
            )
            for h in headers
        ]

        for c, width in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(c + 1)].width = max(width, 8) * 1.2

        # Add the rows
        ws.append(headers)
        for row in table_data:
            ws.append([_jsonify(row.get(header)) for header in headers])

        _add_frontend_formatting(
            ws,
            n_rows=len(table_data) + 1,
            n_cols=len(headers),
            style_set=style_set,
        )

    memory_file = io.BytesIO()
    wb.save(memory_file)
    memory_file.seek(0)

    return memory_file
